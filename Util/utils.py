# -*- coding: utf-8 -*-
# @Time    : 2024/3/28 17:40
# @Author  : yanghangqiao / 杨杭桥
# @Email   : easlie.yang@dbappsecurity.com.cn
# @File    : utils.py.py

import logging
import re
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from Dict.dicts import Dicts


class PublicUtil:

    @staticmethod
    def create_box(text):
        lines = text.split('\n')
        max_length = max(len(line) for line in lines)
        horizontal_line = '+' + '-' * (max_length + 2) + '+'
        boxed_text = [horizontal_line]
        for line in lines:
            boxed_text.append('| ' + line.ljust(max_length) + ' |')
        boxed_text.append(horizontal_line)
        return '\n'.join(boxed_text)

    @staticmethod
    def count_decimal_places(num):
        integer_part, decimal_part = math.modf(num)
        if decimal_part == 0:
            return 0
        else:
            return len(str(decimal_part).split('.')[1])

    @staticmethod
    def create_post_or_put_case(method: str, data: list, param: str, param_type: str):
        case = [Dicts.title[method].format(name=Dicts.API['name'], param=param)]
        for step in range(len(data)):
            val = str(data[step][0])
            if re.match('^大于', val) or re.match('^小于', val):
                value = val
            elif '-' in val and re.match('^(?!-)', val):  # 排除负数
                value = '满足' + val
            else:
                value = '为' + val
            content = Dicts.post_or_put_content['content'].format(step=step + 1,
                                                                  api=Dicts.API['url'],
                                                                  method='新建' if method == 'POST' else '修改',
                                                                  name=Dicts.API['name'],
                                                                  param=param,
                                                                  type=Dicts.param_type[
                                                                      param_type],
                                                                  value=value)
            check = Dicts.post_or_put_content['checks'].format(step=step + 1, result=data[step][1])
            case.extend([content, check])
        return case

    @staticmethod
    def create_get_case(search_key: str, states: list, steps=0):
        case = []
        for step in range(len(states)):
            content = Dicts.get_content['content'].format(step=step + 1 + steps,
                                                          url=Dicts.API['url'], name=Dicts.API['name'], key=search_key,
                                                          state=states[step][0])
            check = Dicts.get_content['checks'].format(step=step + 1 + steps, result=states[step][1])
            case.extend([content, check])
        return case

    @staticmethod
    def create_delete_case(method: str, index_key: str, states: list, batch_api=None, steps=0):
        case = [Dicts.title[method].format(name=Dicts.API['name'])]
        for step in range(len(states)):
            content = Dicts.delete_content['content'].format(step=step + 1 + steps,
                                                             url=Dicts.API['url'] if batch_api is None else batch_api,
                                                             name=Dicts.API['name'], key=index_key,
                                                             state=states[step][0])
            check = Dicts.delete_content['checks'].format(step=step + 1 + steps, result=states[step][1])
            case.extend([content, check])
        return case

    @staticmethod
    def init_xlsx():
        wb = Workbook()
        ws = wb.active
        ws['A1'] = Dicts.excel['A1']
        ws['B1'] = Dicts.excel['B1']
        ws['C1'] = Dicts.excel['C1']
        xlsx_name = f'【{Dicts.API["name"]}】接口测试用例.xlsx'
        wb.save(xlsx_name)
        return xlsx_name

    @staticmethod
    def write_xlsx(xlsx_name: str, cases: list):
        wb = load_workbook(xlsx_name)
        ws = wb.active
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=cases[0])
        # row += 1  # 如需树状展开可解开注释
        for i in range(1, len(cases), 2):
            steps = f'B{row}'
            checks = f'C{row}'
            current_steps = ws[steps].value
            current_checks = ws[checks].value
            ws[steps] = cases[i] if current_steps is None else current_steps + '\n' + cases[i]
            ws[checks] = cases[i + 1] if current_checks is None else current_checks + '\n' + cases[i + 1]
            # ws.cell(row=row, column=2, value=cases[i])    # 如需树状展开可解开注释
            # ws.cell(row=row, column=3, value=cases[i + 1])    # 如需树状展开可解开注释
            # row += 1  # 如需树状展开可解开注释
        wb.save(xlsx_name)


class SelectUtil:

    # 选择用于生成用例的key值
    @staticmethod
    def select_params(keys: list, selections: list):
        print('————\n已识别所有key值如下')
        for index, value in enumerate(keys):
            print(f'{index:2d}' + f': {value}')
        while True:
            idxs = input('请选择需要使用的key值索引并重命名,以-1退出（例如1,2,3,-1）：\n').replace('，', ',').split(',')
            for idx in idxs:
                if int(idx) != -1:
                    if int(idx) < len(keys):
                        selections.append(input(f'{keys[int(idx)].split("—")[-1]}重命名为：'))
                    else:
                        logging.error(f'索引{idx}无效')
                else:
                    print('当前已选择参数')
                    print(selections)
                    return selections

    # 删除已选的key值
    @staticmethod
    def del_params(selections: list):
        while True:
            print('————\n已选择参数如下：')
            for index, value in enumerate(selections):
                print(str(index) + ': ' + value)
            idxs = input('请选择需要删除的参数,以-1退出（例如1,2,3,-1）：\n').replace('，', ',').split(',')
            for idx in idxs:
                if int(idx) != -1:
                    if int(idx) < len(selections):
                        selections.remove(selections[int(idx)])
                        print('删除成功')
                    else:
                        logging.error(f'索引{idx}无效')
                else:
                    print('当前已选择参数')
                    print(selections)
                    return selections

    # 选择测试数据生成方式
    @staticmethod
    def set_post_or_put_methods(selections: list):
        print('————\n分别为已选择的参数选择方法以及对应值：')
        data_util = DataUtil()
        for selection in selections:
            methods = input(f'————\n为字段【{selection}】选择methods\n1-POST  2-PUT  0-exit\n').replace('，', ','). \
                split(',')
            for method in methods:
                if method == '1' or method == '2':
                    if Dicts.excel['name'] == '':
                        xlsx_name = Dicts.excel['name'] = PublicUtil.init_xlsx()
                    else:
                        xlsx_name = Dicts.excel['name']
                    case = int(input('————\n1-等价类  2-边界值  3-自定义  4-bool\n'))
                    if case == 1:
                        param_type = input('请输入参数性质，目前支持：1-长度  2-数值\n')
                        print('————\n请输入合法范围（例如：1-2,5,100-200）')
                        equivalence_classes = data_util.get_values(param_type=param_type, method=1)
                        case = PublicUtil.create_post_or_put_case(method=Dicts.method[method],
                                                                  data=equivalence_classes,
                                                                  param=selection,
                                                                  param_type='length' if param_type == '1' else 'value')
                        PublicUtil.write_xlsx(xlsx_name, case)
                    elif case == 2:
                        param_type = input('请输入参数性质，目前支持：1-长度  2-数值\n')
                        print('————\n请输入有效等价类（例如：1-2,5,100-200）')
                        boundary_values = data_util.get_values(param_type=param_type, method=0)
                        case = PublicUtil.create_post_or_put_case(method=Dicts.method[method], data=boundary_values,
                                                                  param=selection,
                                                                  param_type='length' if param_type == '1' else 'value')
                        PublicUtil.write_xlsx(xlsx_name, case)
                    elif case == 3:
                        valid_custom = input('请输入有效自定义值，例如：预定义服务组,自定义服务组\n').replace('，', ','). \
                            split(',')
                        invalid_custom = input('请输入无效自定义值，例如：不存在的服务组,空\n').replace('，', ',').split(',')
                        custom = []
                        for value in valid_custom:
                            custom.append([value, True])
                        for value in invalid_custom:
                            custom.append([value, False])
                        case = PublicUtil.create_post_or_put_case(method=Dicts.method[method], data=custom,
                                                                  param=selection,
                                                                  param_type='custom')
                        PublicUtil.write_xlsx(xlsx_name, case)
                    else:
                        case = PublicUtil.create_post_or_put_case(method=Dicts.method[method],
                                                                  data=[[True, True], [False, False]],
                                                                  param=selection, param_type='custom')
                        PublicUtil.write_xlsx(xlsx_name, case)

    @staticmethod
    def set_get_or_delete_methods():
        print('————\n分别为已选择的参数选择方法以及对应值：')
        data_util = DataUtil()
        methods = input(f'————\n为字段【{Dicts.API["name"]}】选择methods\n1-GET  2-DELETE  0-exit\n').replace('，', ','). \
            split(',')
        for method in methods:
            if method == '1':
                if Dicts.excel['name'] == '':
                    xlsx_name = Dicts.excel['name'] = PublicUtil.init_xlsx()
                else:
                    xlsx_name = Dicts.excel['name']
                search_keys = input('请输入查询请求的字段，例如：page,size,policy_id\n').replace('，', ',').split(',')
                case = []
                steps = 0
                for index in range(len(search_keys)):
                    states = []
                    valid_states = input(f'请输入字段{search_keys[index]}的有效值，例如：正整数\n').replace('，', ',')\
                        .split(',')
                    invalid_states = input(f'请输入字段{search_keys[index]}的无效值，例如：负数,浮点数\n').replace('，', ',')\
                        .split(',')
                    for state in valid_states:
                        states.append([state, True])
                    for state in invalid_states:
                        states.append([state, False])
                    case.append(PublicUtil.create_get_case(search_key=search_keys[index], states=states, steps=steps))
                    steps += len(states)
                case = [item for sublist in case for item in sublist]
                case.insert(0, Dicts.title['GET'].format(url=Dicts.API['url'], param=Dicts.API['name']))
                print(case)
                PublicUtil.write_xlsx(xlsx_name, case)
            elif method == '2':
                if Dicts.excel['name'] == '':
                    xlsx_name = Dicts.excel['name'] = PublicUtil.init_xlsx()
                else:
                    xlsx_name = Dicts.excel['name']
                index_key = input('请输入删除请求的字段，例如：id\n')
                states = []
                valid_states = input('请输入字段的有效值，例如：存在的id\n').replace('，', ',').split(',')
                invalid_states = input('请输入字段的无效值，例如：不存在的id,被引用的id\n').replace('，', ',').split(',')
                for state in valid_states:
                    states.append([state, True])
                for state in invalid_states:
                    states.append([state, False])
                case = PublicUtil.create_delete_case(method=Dicts.method['4'], index_key=index_key, states=states)
                batch = int(input('是否需要添加批量删除步骤？\n1-需要  0-跳过\n'))
                if batch:
                    batch_api = input('请输入批量删除的api_url\n') + '批量'
                    batch_case = PublicUtil.create_delete_case(method=Dicts.method['4'], index_key=index_key,
                                                               states=states, batch_api=batch_api,
                                                               steps=int((len(case) - 1) / 2))
                    case = case + batch_case[1:]
                print(case)
                PublicUtil.write_xlsx(xlsx_name, case)


class DataUtil:

    # 获取JSON中所有key值
    def get_all_keys(self, json_data, indent=0, keys_list=None):
        if keys_list is None:
            keys_list = []
        # 处理内嵌dict
        if isinstance(json_data, dict):
            for key, value in json_data.items():
                keys_list.append(' ' * indent + '├——' * bool(indent) + str(key))
                self.get_all_keys(value, indent + 2, keys_list)
        # 处理内嵌list
        if isinstance(json_data, list):
            for item in json_data:
                self.get_all_keys(item, indent + 2, keys_list)
        return keys_list

    def get_values(self, param_type: str, method=1):
        borders = input().replace('，', ',').split(',')
        interval_list = [list(map(float, item.split('-')))
                         if '-' in item else [float(item), float(item)]
                         for item in borders]
        merged_interval = []
        for start, end in interval_list:
            merged = False
            for i in range(len(merged_interval)):
                if start <= merged_interval[i][1] and end >= merged_interval[i][0]:
                    merged_interval[i][0] = min(merged_interval[i][0], start)
                    merged_interval[i][1] = max(merged_interval[i][1], end)
                    merged = True
                    logging.warning('请不要试图找bug...')
                    break
            if not merged:
                merged_interval.append([start, end])
        increase = list(filter(lambda x: x[0] <= x[1], merged_interval))  # 排除左值大于右值
        increase = [[int(item) for item in sublist] for sublist in increase] if param_type == '1' else increase
        final_borders = sorted(increase, key=lambda x: int(x[0]) if param_type == '1' else x[0])  # 排序
        if method:
            return self.equivalent(final_borders)
        else:
            return self.border(final_borders)

    # 根据给出的范围确定边界值
    @staticmethod
    def border(borders: list):
        boundary_values = []
        for start, end in borders:
            if start != end:
                start_left = [start - 0.1 ** (PublicUtil.count_decimal_places(start)), False]
                start_right = [start + 0.1 ** (PublicUtil.count_decimal_places(start)), True]
                end_left = [end - 0.1 ** (PublicUtil.count_decimal_places(start)), True]
                end_right = [end + 0.1 ** (PublicUtil.count_decimal_places(start)), False]
                middle = [(start + end) / 2, True]
                boundary_values.extend([start_left, [start, True], start_right,
                                        middle, end_left, [end, True], end_right])
            else:
                boundary_values.append([start, True])
        logging.info('已确认范围' + str(borders))
        logging.info('已确认边界值' + str(boundary_values))
        return boundary_values

    # 根据给出的范围确定等价类
    @staticmethod
    def equivalent(borders: list):
        # total = input('请输入总范围（例如1-100），如果输入0则认为无限大范围').split('-')
        left = borders[0][0]
        right = borders[len(borders) - 1][1]
        # equivalence_classes = [[f'小于{left}', False]] if len(total) == 0 else [[f'{total}']]
        equivalence_classes = [[f'小于{left}', False]]
        for index in range(len(borders)):
            if borders[index][0] != borders[index][1]:
                equivalence_classes.append([f'{borders[index][0]}-{borders[index][1]}', True])
            else:
                equivalence_classes.append([f'{borders[index][0]}', True])
            if index < len(borders) - 1:
                equivalence_classes.append([f'{borders[index][1]}-{borders[index + 1][0]}', False])
        equivalence_classes.append([f'大于{right}', False])
        logging.info('已确认范围' + str(borders))
        logging.info('已确认等价类' + str(equivalence_classes))
        return equivalence_classes
